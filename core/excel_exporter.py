from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side


class ExcelExporter:
    @staticmethod
    def export_quality_report(bubble_manager, save_path):
        """导出质检报告Excel
        列：序号、页码、尺寸、公差、备注
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "质检标注清单"

        # 表头
        headers = ["序号", "页码", "尺寸", "公差", "备注"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 填充数据
        all_bubbles = bubble_manager.get_all_bubbles()
        # 按页码、序号排序
        all_bubbles.sort(key=lambda b: (b["page_num"], int(b["text"]) if b["text"].isdigit() else 0))

        for row_idx, bubble in enumerate(all_bubbles, 2):
            ws.cell(row=row_idx, column=1, value=bubble["text"])
            ws.cell(row=row_idx, column=2, value=bubble["page_num"] + 1)
            ws.cell(row=row_idx, column=3, value=bubble["dimension"])
            ws.cell(row=row_idx, column=4, value=bubble["tolerance"])
            ws.cell(row=row_idx, column=5, value=bubble["remark"])

            # 给所有单元格加边框和居中
            for col in range(1, 6):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # 自动调整列宽
        column_widths = [10, 10, 20, 20, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width

        wb.save(save_path)
        return True